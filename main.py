import openpyxl


def parse(p):
    book = openpyxl.open(p, read_only=True)
    b = openpyxl.load_workbook(filename=path).active
    s = book.active
    print(s.cell(75, 2))
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    week1 = {a: [[None, None]for j in range(6)] for a in days}
    week2 = {a: [[None, None]for j in range(6)] for a in days}
    k = 4
    for day in week1:
        for i in range(6):
            week1[day][i] = [s['B' + str(k)].value, s['C' + str(k)].value]
            k += 2
    k = 4
    for day in week2:
        for i in range(6):
            if s['C' + str(k + 1)].value is not None:
                week2[day][i] = [s['B' + str(k)].value, s['C' + str(k + 1)].value]
                k += 2
            else:
                try:
                    """
                    В общем тут у нас костыль, потому что я не нашел функцию, позволяющую понять объединены ли клетки,
                    поэтому мы проверяем какую ошибку даст функция unmerge_cells, потому что она возвращает разные
                    ошибки при попытке разъединить необъединённые клетки и при попытке разъединить объединённые клетки.
                    Нужен этот костыль голимый для того, чтобы можно было корректно распарсить расписание на вторую
                    неделю, т.к. все значения в объединенный клетках лежали в левой верхней клетке при разъединении
                    """
                    cell_coord = 'C' + str(k) + ':C' + str(k + 1)
                    b.unmerge_cells(cell_coord)
                except ValueError as err:
                    if str(err) == "list.remove(x): x not in list":
                        week2[day][i] = [s['B' + str(k)].value, s['C' + str(k)].value]
                        k += 2
                    else:
                        week2[day][i] = [s['B' + str(k)].value, None]
                        k += 2

    return week1, week2


if __name__ == '__main__':
    path = r'C:\Users\Dooplik\Desktop\raspisanie.xlsx'
    first_week, second_week = parse(path)
    for i in second_week:
        print(i)
        for j in second_week[i]:
            print(j)
        print('-------------------------')

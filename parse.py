import openpyxl
from bs4 import BeautifulSoup
import json
import requests
import time


def parse_xlsx(p, d):
    book = openpyxl.open(p, read_only=True)
    b = openpyxl.load_workbook(filename=p).active
    s = book.active
    days = ['0', '1', '2', '3', '4', '5']
    week1 = {a: [[None, None] in range(6)] for a in days}
    week2 = {a: [[None, None] in range(6)] for a in days}
    k = 4
    for day in week1:
        for i in range(6):
            week1[day][i] = [s['B' + str(k)].value, s['C' + str(k)].value]
            k += 2
    k = 4
    for day in week2:
        for i in range(6):
            if s['C' + str(k + 1)].value is not None:
                week2[day][i] = [s['B' + str(k)].value, s['C' + str(k + 1)].value]
                k += 2
            else:
                try:
                    """
                    В общем тут у нас костыль, потому что я не нашел функцию, позволяющую понять объединены ли клетки,
                    поэтому мы проверяем какую ошибку даст функция unmerge_cells, потому что она возвращает разные
                    ошибки при попытке разъединить необъединённые клетки и при попытке разъединить объединённые клетки.
                    Нужен этот костыль голимый для того, чтобы можно было корректно распарсить расписание на вторую
                    неделю, т.к. все значения в объединенный клетках лежали в левой верхней клетке при разъединении
                    """
                    cell_coord = 'C' + str(k) + ':C' + str(k + 1)
                    b.unmerge_cells(cell_coord)
                except ValueError as err:
                    if str(err) == "list.remove(x): x not in list":
                        week2[day][i] = [s['B' + str(k)].value, s['C' + str(k)].value]
                        k += 2
                    else:
                        week2[day][i] = [s['B' + str(k)].value, None]
                        k += 2

    result = ''
    for i in week1[str(d)]:
        if i[1] is not None:
            i[1] = i[1].replace('\n', ' ')
        result += f'{i[0]} || {i[1]}\n'
    return result


def parse_teachers():
    url = 'https://pstu.ru/title1/faculties/'
    headers = {
        'Accept': '*/*',
        'User-Agent': 'Mozilla / 5.0(Windows NT 10.0;Win64;x64) AppleWebKit /'
                      ' 537.36(KHTML, likeGecko) Chrome / 102.0.0.0Safari / 537.36'
    }
    req = requests.get(url, headers=headers)
    src = req.text
    soup = BeautifulSoup(src, 'lxml')
    faculties = soup.find('li', text='Кафедра Электротехника и электромеханика (ЭТЭМ)') \
        .find_parent() \
        .find_parent() \
        .find_parent()
    with open('departments.json') as f:
        loaded_json = json.load(f)
    a = 'AKF'
    for i in faculties.find_all('a'):
        if 'базовая' not in i.text.lower():
            if 'иностранных студентов' in i.text:
                break
            elif 'факультет' in i.text.lower():
                a = i.get('href').split('/')[-2].upper()
                loaded_json['faculties'][i.get('href').split('/')[-2].upper()] = {}
            if '(' and ')' in i.text:
                url = f'https://pstu.ru{i.get("href")}'
                req = requests.get(url, headers=headers)
                src = req.text
                soup = BeautifulSoup(src, 'lxml')
                link = soup.find("a", text="Сотрудники кафедры")
                try:
                    loaded_json['faculties'][a][i.text] = f'https://pstu.ru{link.get("href")}'
                except AttributeError:
                    print(i)

    with open('teachers.json') as f:
        teachers_json = json.load(f)
    for fac in loaded_json['faculties']:
        for dep in loaded_json['faculties'][fac]:
            url = loaded_json['faculties'][fac][dep]
            req = requests.get(url, headers=headers)
            src = req.text
            soup = BeautifulSoup(src, 'lxml')
            teach_tag = soup.find("div", class_="fac").find_all("a")
            teach_list = [[teach.text, f"https://pstu.ru{teach.get('href')}"] for teach in teach_tag]
            teachers_json['faculties'][fac][dep] = {}
            for teach in teach_list:
                teachers_json['faculties'][fac][dep][teach[0]] = teach[1]
    with open('teachers.json', 'w') as f:
        json.dump(teachers_json, f, indent=2, ensure_ascii=False)


def test():
    with open('index.html', encoding='UTF-8') as file:
        src = file.read()
    soup = BeautifulSoup(src, 'lxml')
    teach_tag = soup.find("div", class_="fac").find_all("a")
    teach_list = [[teach.text, f"https://pstu.ru{teach.get('href')}"] for teach in teach_tag]
    with open('departments.json') as f:
        loaded_json = json.load(f)
    for i in teach_list:
        loaded_json['faculties']['FPMM'][i[0]] = i[1]
    with open('departments.json', 'w') as f:
        json.dump(loaded_json, f, indent=2, ensure_ascii=False)


def parse_groups():
    url = 'https://pstu.ru/student/new_timetable/'
    headers = {
        'Accept': '*/*',
        'User-Agent': 'Mozilla / 5.0(Windows NT 10.0;Win64;x64) AppleWebKit /'
                      ' 537.36(KHTML, likeGecko) Chrome / 102.0.0.0Safari / 537.36'
    }
    req = requests.get(url, headers=headers)
    src = req.text
    with open('groups.json') as f:
        groups_json = json.load(f)
    groups_json['faculties'] = {}
    soup = BeautifulSoup(src, 'lxml')
    for fac in soup.find_all('div', style='padding:3px 5px 3px 5px'):
        groups_json['faculties'][fac.text] = {}
        url = 'https://pstu.ru/student/new_timetable/' + fac.find('a').get('href')
        req = requests.get(url, headers=headers)
        src = req.text
        soup = BeautifulSoup(src, 'lxml')
        for group_list in soup.find_all('div', style='padding:3px 5px 3px 25px'):
            groups_json['faculties'][fac.text][group_list.text] = {}
            url = 'https://pstu.ru/student/new_timetable/' + group_list.find('a').get('href')
            req = requests.get(url, headers=headers)
            src = req.text
            soup = BeautifulSoup(src, 'lxml')
            for group in soup.find_all('div', style='padding:3px 5px 3px 45px'):
                groups_json['faculties'][fac.text][group_list.text][group.text.replace(' ', '')] = {}
                url = 'https://pstu.ru/student/new_timetable/' + group.find('a').get('href')
                req = requests.get(url, headers=headers)
                src = req.text
                soup = BeautifulSoup(src, 'lxml')
                for rasp in soup.find_all('div', style='padding:3px 5px 3px 65px'):
                    groups_json['faculties'][fac.text][group_list.text][group.text.replace(' ', '')][rasp.text] =\
                        rasp.find('a').get('href')

    with open('groups.json', 'w') as f:
        json.dump(groups_json, f, indent=2, ensure_ascii=False)


if __name__ == '__main__':
    start_time = time.time()
    parse_groups()
    print("%s секунд" % (time.time() - start_time))
